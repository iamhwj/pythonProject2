from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


def style_modify(filename=r"D:\PycharmProjects\KeNuo\2023-04-01\2023-04-01直送.xlsx"):
    wb = load_workbook(filename)  # ex_a.xlsx是文件名
    print(wb.sheetnames)
    for i in wb.sheetnames:
        ws = wb[i]
        # 对O列添加自动换行
        for row in ws['O']:
            row.alignment = Alignment(wrap_text=True)
        for row in ws['G']:
            row.alignment = Alignment(wrap_text=True)
        for cell in ws[1]:
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.rows:
            for cell in row:
                cell.alignment = cell_alignment
        ws.row_dimensions[1].height = 30
        # 设置每列的宽度
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
    wb.save(filename)


if __name__ == '__main__':
    style_modify()
